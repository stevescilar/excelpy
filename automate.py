from openpyxl import Workbook, load_workbook

wb = load_workbook('marks.xlsx')
ws = wb.active
# print(ws) -- print active sheet
# print(ws['b2'].value)
# ws['B3'].value = 'Kithungo Raha'

# wb.save('marks.xlsx')
# print(wb.sheetnames)
# create a new sheet 
wb.create_sheet('Manuel')
wb.save('marks.xlsx')
print(wb.sheetnames)