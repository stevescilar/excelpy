from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# create a workbook object

# wb = Workbook()


# load existing excel sheet

wb = load_workbook('marks.xlsx')

# create an active worksheet
ws = wb.active

# # set a variable
# name = ws["B3"].value
# reg = ws["c3"].value
# #print something from loaded spreadsheat
# # print(ws["B3"].value)
# print(f'{name}:{reg}')

# grab a whole column
# col_a = ws['2']
# # print(col_a)
# # sort a tupple
# for cell in col_a:
#     print(cell.value)

# grab range
range = ws['b3':'b9']

for cell in range:
    for x in cell:
        print(x.value)